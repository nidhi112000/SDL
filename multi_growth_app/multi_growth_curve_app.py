#!/usr/bin/env python3

import logging
from argparse import ArgumentParser
import time
from tools.gladier_flow.growth_curve_gladier_flow import c2_flow
from pathlib import Path
from tools.hudson_solo_auxillary.hso_functions import package_hso
from tools.hudson_solo_auxillary import solo_multi_step1, solo_multi_step2, solo_multi_step3
import pandas as pd 
import pathlib
import openpyxl
import numpy as np
import os
import datetime
import random
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from tools.ai_model import ai_actions
import scipy.stats as stats

from rpl_wei import Experiment

#from rpl_wei.wei_workcell_base import WEI

ORIGINAL_ANTIBIOTIC_CONCENTRATION = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
ORIGINAL_CELL_CONCENTRATION = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
TOTAL_CELL_COLUMN_CONCENTRATION = [] #Each row represents another plate, each value is the concentration of the cell in the column of the plate
TOTAL_TREATMENT_COLUMN_CONCENTRATION = [] #Each row represents another plate, each value is the concentration of the cell in the column of the plate
CULTURE_PAYLOAD = []
MEDIA_PAYLOAD = []
HIDEX_UPLOADS = []
COMPLETED_CELL_COLUMNS = []
COMPLETED_ANTIBIOTIC_COLUMNS = []
PLATE_BARCODES = []
CREATED_COMPLETED_FILE = False
COMPLETED_FILE_NAME = ''
EXPERIMENT_FILE_PATH = ''

COMPLETE_HUDSON_SETUP_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/complete_hudson_setup.yaml'
STREAMLINED_HUDSON_SETUP_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/streamlined_hudson_setup.yaml'
SETUP_GROWTH_MEDIA_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/setup_growth_media.yaml'

CREATE_PLATE_T0_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/create_plate_T0.yaml'
READ_PLATE_T12_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/read_plate_T12.yaml'

DISPOSE_BOX_PLATE_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/dispose_box_plate.yaml'
DISPOSE_GROWTH_MEDIA_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/dispose_growth_media.yaml'

OPEN_CLOSE_HIDEX_FILE_PATH = '/home/rpl/workspace/BIO_workcell/multi_growth_app/workflows/open_close_hidex.yaml'

exp = Experiment('127.0.0.1', '8000', 'Growth_Curve')
exp.register_exp() 
exp.events.log_local_compute("package_hso")

def sample_method_implementing_ai():
    ai_actions.load_model()
    path_name = str(pathlib.Path().resolve()) + "\\multi_growth_app\\completed_runs\\" + "07-20-2023 at 13.19.40 Completed Run.xlsx"
    ai_actions.train_model(path_name)
    treatment_column, treatment_concentration, cell_column, cell_concentration, predictions = ai_actions.predict_experiment(3 ,ORIGINAL_ANTIBIOTIC_CONCENTRATION, ORIGINAL_CELL_CONCENTRATION)
    ai_actions.save_model()

#Data Processing Functions
def process_experimental_results():
    global HIDEX_UPLOADS
    global COMPLETED_CELL_COLUMNS
    global COMPLETED_ANTIBIOTIC_COLUMNS
    global PLATE_BARCODES
    global CREATED_COMPLETED_FILE
    global COMPLETED_FILE_NAME
    global CULTURE_PAYLOAD
    global MEDIA_PAYLOAD

    globus_runs_df = pd.DataFrame(columns=['Plate #', 'Well', 'Reading Hour', 'Result'])
    #Here, we are uploading all of the
    for upload_id in HIDEX_UPLOADS:
        is_t0_run = False
        run_number = 0
        if upload_id.startswith("T0_") :
            is_t0_run = True
            run_number = int(upload_id[11:-9])
        elif upload_id.startswith("T12_"):
            is_t0_run = False
            run_number = int(upload_id[12:-9])
        single_reading_df = read_globus_data(title_name = upload_id, t0_reading = is_t0_run, plate_number=run_number)
        globus_runs_df = pd.concat([globus_runs_df, single_reading_df], ignore_index=True)

    print(globus_runs_df)

    old_t0_run_ids = globus_runs_df['Plate #'].drop_duplicates().values
    old_t0_run_ids = old_t0_run_ids.astype(int)

    globus_runs_df['Plate #'] = globus_runs_df['Plate #'].astype(int)
    filter_check_runs_df = globus_runs_df.copy()
    filter_check_runs_df.sort_values(by=['Plate #', 'Well'], inplace=True)
    filter_check_runs_df.reset_index(drop=True, inplace=True)

    #Filtering runs without a t0 or t12 variable
    plates_to_remove = []
    prev_start_time = None
    for index, row in filter_check_runs_df.iterrows():
        current_start_time = row['Reading Hour']
        if prev_start_time is None:
            prev_start_time = current_start_time
        elif prev_start_time == current_start_time:
            plates_to_remove.append(row['Plate #'])
        else:
            prev_start_time = current_start_time
    filtered_globus_runs_df = globus_runs_df[~globus_runs_df['Plate #'].isin(plates_to_remove)]

    print(filtered_globus_runs_df)

    filtered_t0_run_ids = filtered_globus_runs_df['Plate #'].drop_duplicates().values
    filtered_t0_run_ids = filtered_t0_run_ids.astype(int)

    removed_plate_numbers = set(plates_to_remove)
    num_removed_plates = len(removed_plate_numbers)

    cell_columns = []
    antibiotic_string_columns = []
    barcodes = []
    
    for run_id in filtered_t0_run_ids:
        info_index = old_t0_run_ids.tolist().index(run_id)
        cell_columns.append(COMPLETED_CELL_COLUMNS[info_index])
        antibiotic_string_columns.append(COMPLETED_ANTIBIOTIC_COLUMNS[info_index])
        barcodes.append(PLATE_BARCODES[info_index])

    antibiotic_columns = []
    for string_column in antibiotic_string_columns:
        integer_value = int(string_column[3:])
        antibiotic_columns.append(integer_value)

    print(antibiotic_columns)
    print(cell_columns)

    antibiotic_concentrations_list = []
    for antibiotic_column in antibiotic_columns:
        antibioitic_index = antibiotic_column - 1
        single_column_antibiotic_concentration_list = TOTAL_TREATMENT_COLUMN_CONCENTRATION[antibioitic_index]
        single_plate_all_antibiotic_concentrations = []
        for i in range(0,8):
            for j in range(0,12):
                single_plate_all_antibiotic_concentrations.append(single_column_antibiotic_concentration_list[j])
        antibiotic_concentrations_list.append(single_plate_all_antibiotic_concentrations)

    cell_concentrations_list = []
    for cell_column in cell_columns:
        cell_index = cell_column - 1
        single_column_cell_concentration_list = TOTAL_CELL_COLUMN_CONCENTRATION[cell_index]
        single_plate_all_cell_concentrations = []
        for i in range(0,8):
            for j in range(0,12):
                single_plate_all_cell_concentrations.append(single_column_cell_concentration_list[j])
        cell_concentrations_list.append(single_plate_all_cell_concentrations)

    print(antibiotic_concentrations_list)
    print(cell_concentrations_list)

    folder_path = str(pathlib.Path().resolve()) + "/completed_runs/"
    #folder_path = str(pathlib.Path().resolve()) + "\\multi_growth_app\\completed_runs\\"
    current_sheet_index = 1

    if CREATED_COMPLETED_FILE:
        print("Completed File Name ", COMPLETED_FILE_NAME)
        path_name = folder_path + COMPLETED_FILE_NAME
        completed_workbook = openpyxl.load_workbook(path_name)
        num_sheets = len(completed_workbook.worksheets)
        current_sheet_index = num_sheets + 1

    else:
        print("Creating Excel Object")
        current_date = datetime.date.today()
        formatted_date = current_date.strftime("%m-%d-%Y")
        formatted_time = str(time.strftime("%H.%M.%S", time.localtime()))
        file_name = formatted_date + " at " + formatted_time +  " Completed Run" + ".xlsx"
        COMPLETED_FILE_NAME = file_name
        completed_workbook = openpyxl.Workbook() 
        CREATED_COMPLETED_FILE = True
        os.makedirs(folder_path, exist_ok=True)
        path_name = folder_path + COMPLETED_FILE_NAME
        completed_workbook.save(os.path.join(path_name))
        default_sheet = completed_workbook.active
        completed_workbook.remove(default_sheet)

    for i in range(0, len(filtered_t0_run_ids)):
        sheet_name = "Run " + str(int(i + 1))
        current_sheet = completed_workbook.create_sheet(sheet_name)
        t0_specific_run_df = filtered_globus_runs_df[(globus_runs_df['Plate #'] == filtered_t0_run_ids[i]) & (globus_runs_df['Reading Hour'] == 'T0')].copy()
        t0_specific_run_df.reset_index(drop=True, inplace=True)
        print(t0_specific_run_df)
        t12_specific_run_df = filtered_globus_runs_df[(globus_runs_df['Plate #'] == filtered_t0_run_ids[i]) & (globus_runs_df['Reading Hour'] == 'T12')].copy()
        t12_specific_run_df.reset_index(drop=True, inplace=True)
        print(t12_specific_run_df)
        print("antibiotic column ", antibiotic_columns)
        runs_df = pd.DataFrame(columns=['Well', 'Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration', 'Growth Rate', 'T0 Reading', 'T12 Reading'])
        control_df = pd.DataFrame(columns=['Well', 'Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration', 'Blank Growth Rate', 'Blank T0 Reading', 'Blank T12 Reading'])
        for j in range(0,96):
            t0_growth_value = float(t0_specific_run_df.loc[j, 'Result'])
            t12_growth_value = float(t12_specific_run_df.loc[j, 'Result'])
            growth_rate = t12_growth_value - t0_growth_value
            well_index = chr(65 + int((j - j % 12)/12)) + str(int(j % 12 + 1))
            if((j-j%12)/12)%2 == 0:
                latest_row = {
                    'Well' : well_index,
                    'Treatment Column': antibiotic_columns[i], 
                    'Treatment Concentration': antibiotic_concentrations_list[i][j],
                    'Cell Column': cell_columns[i],
                    'Cell Concentration': cell_concentrations_list[i][j],
                    'Growth Rate' : growth_rate,
                    'T0 Reading' : t0_growth_value,
                    'T12 Reading' : t12_growth_value
                }
                latest_row_df = pd.DataFrame(latest_row, index=[0])
                runs_df = pd.concat([runs_df, latest_row_df], ignore_index=True)
            else:
                blank_latest_row = {
                    'Well' : well_index,
                    'Treatment Column': 0.0, 
                    'Treatment Concentration': 0.0,
                    'Cell Column': 0.0,
                    'Cell Concentration': 0.0,
                    'Blank Growth Rate' : growth_rate,
                    'Blank T0 Reading' : t0_growth_value,
                    'Blank T12 Reading' : t12_growth_value
                }
                blank_latest_row = pd.DataFrame(blank_latest_row, index=[0])
                control_df = pd.concat([control_df, blank_latest_row], ignore_index=True)
        
        for _ in range (0,3):
            current_sheet.append([])
    
        for row in dataframe_to_rows(runs_df, index=False, header=True):
            current_sheet.append(row)
        
        current_sheet.append([])

        for row in dataframe_to_rows(control_df, index=False, header=True):
            current_sheet.append(row)

        current_sheet['A1'] = "Barcode Number"
        current_sheet['B1'] = barcodes[i]

        current_sheet['A2'] = "Slope (Best-Fit Line)"
        slope,intercept = return_line_of_best_fit(current_sheet)
        current_sheet['B2'] = slope
        current_sheet['C2'] = "Y-Intercept (Best-Fit Line)"
        current_sheet['D2'] = intercept

    save_path = folder_path + COMPLETED_FILE_NAME
    completed_workbook.save(save_path)

    CULTURE_PAYLOAD = []
    MEDIA_PAYLOAD = []
    HIDEX_UPLOADS = []
    COMPLETED_CELL_COLUMNS = []
    COMPLETED_ANTIBIOTIC_COLUMNS = []
    PLATE_BARCODES = []

def read_globus_data(title_name = '', t0_reading = True, plate_number = 0):
    print("Reading Globus Data -- Accessed Chrome")
    driver = webdriver.Chrome()
    driver.get("https://acdc.alcf.anl.gov/sdl-bio/?q=*")
    search_bar = driver.find_element(By.ID, "search-input")
    search_query = "\"" + title_name + "\""
    search_bar.send_keys(search_query)
    search_bar.submit()

    link_element = driver.find_element(By.LINK_TEXT, title_name)
    link_element.click()

    parent_element = driver.find_element("css selector", ".col-md-6")
    table_elements = parent_element.find_elements(By.TAG_NAME, "table")
    desired_table_element = table_elements[1]

    rows = desired_table_element.find_elements(By.TAG_NAME, "tr")
    table_data = []
    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        row_data = [cell.text for cell in cells]
        table_data.append(row_data)
    
    driver.quit()

    # globus_df = pd.DataFrame(columns=['Plate #', 'Well', 'Reading Hour'])
    # plate_array = [plate_number] * 96
    # globus_df['Plate #'] = plate_array
    # if t0_reading == True:
    #     times_array = ["T0"] * 96
    #     globus_df['Reading Hour'] = times_array
    # else: 
    #     times_array = ["T12"] * 96
    #     globus_df['Reading Hour'] = times_array
    # well_indices =[]
    # for i in range (0, 96):
    #     well_index = chr(65 + int((i - i % 12)/12)) + str(int(i % 12 + 1))
    #     well_indices.append(well_index)
    # globus_df["Well"] = well_indices
    # results_list = table_data[0]
    # results_list[0] = results_list[0][1:]
    # results_list[len(results_list)-1] = results_list[len(results_list)-1][:-1]
    # globus_df['Result'] = results_list


    globus_df = pd.DataFrame(table_data)
    globus_df.iloc[0][3] = 'Result'
    globus_df.iloc[0][2] = 'Reading Hour'
    globus_df.columns = globus_df.iloc[0]
    globus_df = globus_df[1:]
    globus_df = globus_df.reset_index(drop=True)
    globus_df['Plate #'] = plate_number
    if t0_reading == True:
        globus_df.iloc[:, 2] = "T0"
    else: 
        globus_df.iloc[:, 2] = "T12"

    return globus_df

def return_line_of_best_fit(worksheet):
    x_data = []
    y_data = []
    for i in range(5,53):
        x_data_index = "C" + str(int(i))
        x_data.append(worksheet[x_data_index].value)

        y_data_index = "F" + str(int(i))
        y_data.append(worksheet[y_data_index].value)
    slope, intercept, r_value, p_value, std_err = stats.linregress(x_data, y_data)
    return slope, intercept

def delete_experiment_excel_file():
    global EXPERIMENT_FILE_PATH
    os.remove(EXPERIMENT_FILE_PATH)
    print(EXPERIMENT_FILE_PATH)

def determine_payload_from_excel():
    global EXPERIMENT_FILE_PATH

    print("Run Log Starts Now")
    #folder_path = str(pathlib.Path().resolve()) + "\\multi_growth_app\\active_runs"
    folder_path = str(pathlib.Path().resolve()) + "/active_runs"
    files = os.listdir(folder_path)
    excel_files = [file for file in files if file.endswith(".xlsx")]
    sorted_files = sorted(excel_files, key=lambda x: os.path.getmtime(os.path.join(folder_path, x)))
    path_name = os.path.join(folder_path, sorted_files[0])
    EXPERIMENT_FILE_PATH = path_name
    print(path_name)
    workbook = openpyxl.load_workbook(filename=path_name)
    worksheet = workbook['Complete_Run_Layout']
    experiment_iterations = worksheet['B1'].value
    incubation_time_hours = worksheet['B2'].value
    incubation_time_seconds = incubation_time_hours * 3600
    added_items = 0
    for i in range(2,14):
        column_letter = chr(ord('@')+i)
        run_number_cell_id = column_letter + "3"
        media_type_cell_id = column_letter + "4"
        culture_type_cell_id = column_letter + "5"
        if(worksheet[run_number_cell_id].value != None and worksheet[media_type_cell_id].value != None and worksheet[culture_type_cell_id].value != None):      
            MEDIA_PAYLOAD.append(worksheet[media_type_cell_id].value)
            CULTURE_PAYLOAD.append(worksheet[culture_type_cell_id].value) 
            added_items = added_items + 1
    if(len(MEDIA_PAYLOAD) != experiment_iterations):
        experiment_iterations = len(MEDIA_PAYLOAD)

    for i in range(0,12):
        original_concentration = ORIGINAL_ANTIBIOTIC_CONCENTRATION[i]
        single_plate_treatment_columns = []
        for iterations in range(0,2):
            for j in range (1,6):
                single_plate_treatment_columns.append(original_concentration/2**j)
            single_plate_treatment_columns.append(0)
        TOTAL_TREATMENT_COLUMN_CONCENTRATION.append(single_plate_treatment_columns)

    for i in range(0,12):
        original_concentration = ORIGINAL_CELL_CONCENTRATION[i]
        single_plate_cell_columns = []
        for iterations in range(0,12):
            single_plate_cell_columns.append(original_concentration/10)
        TOTAL_CELL_COLUMN_CONCENTRATION.append(single_plate_cell_columns)
    
    print(CULTURE_PAYLOAD)
    print(MEDIA_PAYLOAD)
    
    return experiment_iterations, incubation_time_seconds

#BIO Workflow Functions
def run_experiment(total_iterations, incubation_time_sec): 
    iterations = 0
    removals = 0
    incubation_start_times = [] 
    print("Total Experimental Runs ", total_iterations)
    print("Current Iteration Variable: ", iterations)
    print("Incubation Time (Seconds): ", incubation_time_sec)
    # The experiment will run until all of the plates are used (indicated by iterations < total_iterations) and there are no more plates in the incubator (indicated by len(incubation_start_times) != 0)
    while(iterations < total_iterations or len(incubation_start_times) != 0):
        #Check to see if there are any more plates to run, indicated by total_iterations
        if(iterations < total_iterations):
            #Debug Log
            print("Starting Experiment ", iterations, ": Started Loop")
            #Set up the experiment based on the number of iterations passed.
            # setup(iterations)
            #Calculate the ID of the plate needed for incubation based on the number of iterations that have passed
            liconic_id = iterations + 1
            #Run the experiment from the Hudson Solo step to the incubation step at a specified Liconic ID
            print("Starting T0 Reading")
            # T0_Reading(liconic_id)
            print("Finished T0 Reading")
            assign_barcode()
            #Add the time of the incubation start to the array of 96 well plates that are currently in the incubator
            incubation_start_times.append(round(time.time()))
            #Since an iteration has now passed (the plate is in the incubator), increase the index of incubation iterations
            iterations = iterations + 1
            #Debug Log
            print("Completed Iterations: ", iterations, " ... Start Times: " , incubation_start_times)
            #Based on the total number of completed incubation iterations, determine what needs to be disposed of from the experimental setup.
            if(iterations % 2 == 0):
                print("Starting Disposal")
                # dispose(iterations)
                print("Ending Disposal")
        # Check to see if delta current time and the time at which the well plate currently incubating longest exceeds the incubation time.
        if(round(time.time()) - incubation_start_times[0] > incubation_time_sec):
            #Debug Log
            print("Finishing Up Experiment ", removals, ": Ending Loop")
            #Calcuate the ID of the 96 well plate needed for removal from the incubator based on the number of plates that have already been removed.
            liconic_id = removals + 1
            #Complete the experiment starting from the removal of the 96 well plate at the specified incubation ID and ending at a Hidex T12 Reading.
            print("Starting T12 Reading")
            T12_Reading(liconic_id)
            print("Ending T12 Reading")
            #Remove the incubation start time of the plate that was just read from the array to update the new longest incubating well plate
            incubation_start_times.pop(0)
            #Increase the total number of removals that have now occurred.
            removals = removals + 1

def dispose(completed_iterations):
    #Set the default disposal index of the serial dilution plate to Stack 2.
    disposal_index = "Stack2"
    #To identify the LidNest location of where certain serial dilution plates must be held, divide the completed iterations by 2
    stack_type = completed_iterations/2
    #For the first two serial dilution plates, define the disposal index as LidNest 1 for the first serial dilution plate and LidNest 2 for the second serial dilution plate
    if(stack_type <= 2):
        disposal_index = "LidNest" + str(int(stack_type))

    #For the fourth serial dilution plate, define the disposal index as LidNest 3.
    #We are not defining the disposal index as LidNest 3 for the third serial dilution plate because there will already be a growth media plate on LidNest 3. This growth media plate will be removed in the subsequent setup function
    if(stack_type == 4):
        disposal_index = "LidNest3"
    #Add the disposal index to the payload
    payload={
        'disposal_location':  disposal_index,    
        }
    #Run the despose Yaml File with the dedicated disposal location
    run_WEI(DISPOSE_BOX_PLATE_FILE_PATH, payload)
    if(stack_type % 3 == 0):
    #If the Stack Type is a multiple of 3 (6 complete iterations of the Hudson experiment have occurred), dispose of the growth media deep well plate
        run_WEI(DISPOSE_GROWTH_MEDIA_FILE_PATH, None)

def setup(iteration_number):
    #If currently on an even number of iterations, add a tip box, serial dilution plate, and 96 well plate to the experiment.
    if(iteration_number % 2 == 0):
        #Identify the Tip Box Position Index, so it can be refilled on the Hudson Client
        payload={
                'tip_box_position': "3",
            }
        #Run the Yaml file that outlines the setup procedure for the tip box, serial dilution plate, and 96 well plate.
        print("Starting Complete Hudson Setup")
        run_WEI(COMPLETE_HUDSON_SETUP_FILE_PATH, payload)
        print("Finished Complete Setup")
        #If currently on a number of iterations that is a factor of 6, add a growth media plate to the experiment as well
        if(iteration_number % 6 == 0):
            #Specify the LidNest index of the growth well media plate that will be added to the setup (This equation assumes that there are only two growth media plates on LidNest 2 and LidNest 3 respectively for a total of 12 runs)
            LidNest_index = 2 + iteration_number/6
            #Convert the index to a readable string
            plateCrane_readable_index = "LidNest" + str(int(LidNest_index))
            print("LidNest Being Used: ", plateCrane_readable_index)
            #Add the LidNest Index to the payload
            payload={
                'lidnest_index':  plateCrane_readable_index,
            }
            #Run the Yaml file that outlines the setup procedure for the growth media plate
            print("Starting Growth Media Setup")
            run_WEI(SETUP_GROWTH_MEDIA_FILE_PATH, payload)
            print("Finished Growth Media Setup!")
    #If currently on an even number of iterations, add a 96 well plate to the experiment.
    else: 
        #Run the Yaml file that outlines the setup procedure for ONLY the 96 well plate
        run_WEI(STREAMLINED_HUDSON_SETUP_FILE_PATH, None)

def T0_Reading(liconic_plate_id):
    plate_id = '' + str(int(liconic_plate_id))
    treatment_col_id = "col" + str(int(MEDIA_PAYLOAD[liconic_plate_id-1]))
    culture_col_id = int(CULTURE_PAYLOAD[liconic_plate_id-1])
    treatment_dilution = 1
    if liconic_plate_id % 2 == 0:
        treatment_dilution = 2
    else: 
        treatment_dilution = 1
    payload={
        'temp': 37.0, 
        'humidity': 95.0,
        'shaker_speed': 30,
        "stacker": 1, 
        "slot": 1,
        "treatment": treatment_col_id, # string of treatment name. Ex. "col1", "col2"
        "culture_column": culture_col_id,  # int of cell culture column. Ex. 1, 2, 3, etc.
        "culture_dil_column": liconic_plate_id % 12, # int of dilution column for 1:10 culture dilutions. Ex. 1, 2, 3, etc.
        "media_start_column": (2*liconic_plate_id-1)%12,  # int of column to draw media from (requires 2 columns, 1 means columns 1 and 2) Ex. 1, 3, 5, etc.
        "treatment_dil_half": treatment_dilution,  #  int of which plate half to use for treatment serial dilutions. Options are 1 or 2. 
        "incubation_plate_id" : plate_id,        
        }

    # from somewhere import create_hso? or directly the solo script
    hso_1, hso_1_lines, hso_1_basename = package_hso(solo_multi_step1.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp1.hso") 
    hso_2, hso_2_lines, hso_2_basename = package_hso(solo_multi_step2.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp2.hso")  
    hso_3, hso_3_lines, hso_3_basename = package_hso(solo_multi_step3.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp3.hso")  

    # update payload with solo hso details
    payload['hso_1'] = hso_1
    payload['hso_1_lines'] = hso_1_lines
    payload['hso_1_basename'] = hso_1_basename

    payload['hso_2'] = hso_2
    payload['hso_2_lines'] = hso_2_lines
    payload['hso_2_basename'] = hso_2_basename

    payload['hso_3'] = hso_3
    payload['hso_3_lines'] = hso_3_lines
    payload['hso_3_basename'] = hso_3_basename

    # #run Growth Create Plate
    hidex_upload_id=run_WEI(CREATE_PLATE_T0_FILE_PATH, payload, Hidex_Used=True, Plate_Number=liconic_plate_id)

    HIDEX_UPLOADS.append(hidex_upload_id)
    COMPLETED_ANTIBIOTIC_COLUMNS.append(treatment_col_id)
    COMPLETED_CELL_COLUMNS.append(culture_col_id)

def T12_Reading(liconic_plate_id):
    plate_id = '' + str(int(liconic_plate_id))

    payload={
        'temp': 37.0, 
        'humidity': 95.0,
        'shaker_speed': 30,
        "stacker": 1, 
        "slot": 1,
        "treatment": "col1", # string of treatment name. Ex. "col1", "col2"
        "culture_column": 1,  # int of cell culture column. Ex. 1, 2, 3, etc.
        "culture_dil_column": 1, # int of dilution column for 1:10 culture dilutions. Ex. 1, 2, 3, etc.
        "media_start_column": 1,  # int of column to draw media from (requires 2 columns, 1 means columns 1 and 2) Ex. 1, 3, 5, etc.
        "treatment_dil_half": 1,  #  int of which plate half to use for treatment serial dilutions. Options are 1 or 2.
        "incubation_plate_id" : plate_id,
        }

    # from somewhere import create_hso? or directly the solo script
    hso_1, hso_1_lines, hso_1_basename = package_hso(solo_multi_step1.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp1.hso") 
    hso_2, hso_2_lines, hso_2_basename = package_hso(solo_multi_step2.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp2.hso")  
    hso_3, hso_3_lines, hso_3_basename = package_hso(solo_multi_step3.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp3.hso")  

    # update payload with solo hso details
    payload['hso_1'] = hso_1
    payload['hso_1_lines'] = hso_1_lines
    payload['hso_1_basename'] = hso_1_basename

    payload['hso_2'] = hso_2
    payload['hso_2_lines'] = hso_2_lines
    payload['hso_2_basename'] = hso_2_basename

    payload['hso_3'] = hso_3
    payload['hso_3_lines'] = hso_3_lines
    payload['hso_3_basename'] = hso_3_basename

    # # #run Growth Create Plate
    hidex_upload_id=run_WEI(READ_PLATE_T12_FILE_PATH, payload, Hidex_Used=True, Plate_Number=liconic_plate_id)
    HIDEX_UPLOADS.append(hidex_upload_id)

def run_WEI(file_location, payload_class, Hidex_Used = False, Plate_Number = 0):
    flow_info = exp.run_job(Path(file_location).resolve(), payload=payload_class, simulate=False)

    flow_status = exp.query_job(flow_info["job_id"])
    while(flow_status["status"] != "finished" and flow_status["status"] != "failure"):
        flow_status = exp.query_job(flow_info["job_id"])
        time.sleep(3)

    run_info = flow_status["result"]
    run_info["run_dir"] = Path(run_info["run_dir"])
    print(run_info)

    if Hidex_Used:
        t0_reading = False
        if file_location == CREATE_PLATE_T0_FILE_PATH:
            t0_reading = True
        else:
            t0_reading = False
        hidex_file_path = run_info["hist"]["run Hidex"]["action_msg"]
        hidex_file_path = hidex_file_path.replace('\\', '/')
        hidex_file_path = hidex_file_path.replace("C:/", "/C/")
        flow_title = Path(hidex_file_path) #Path(run_info["hist"]["run_assay"]["step_response"])
        fname = flow_title.name
        flow_title = flow_title.parents[0]
        experiment_time = str(time.strftime("%H_%M_%S", time.localtime()))
        experiment_name = ''
        hour_read = ''
        if t0_reading:
            experiment_name = "T0_Reading"
            hour_read = '0'
        else:
            experiment_name = "T12_Reading"
            hour_read = '12'

        c2_flow(exp_name = experiment_name, plate_n = str(int(Plate_Number)), time = experiment_time, local_path=flow_title, fname = fname, hour=hour_read, exp = exp)
        print("Finished Uplodaing to Globus")
        return experiment_name + '_' + str(int(Plate_Number)) + '_' + experiment_time

def assign_barcode():
    current_barcode = return_barcode()
    PLATE_BARCODES.append(current_barcode)

def return_barcode():
    barcode = ""
    for _ in range(13):
        digit = random.randint(0, 9)
        barcode += str(digit)
    return barcode

#Experiment Run
def main():
    # iteration_runs, incubation_time = determine_payload_from_excel()
    run_experiment(0, 43200)#(iteration_runs, incubation_time)
    try:
        process_experimental_results()
    except Exception as e:
        print("An exception occurred: ", e)
    # delete_experiment_excel_file()

if __name__ == "__main__":
    #main()
    main()

#!/usr/bin/env python3