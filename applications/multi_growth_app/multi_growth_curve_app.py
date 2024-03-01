#!/usr/bin/env python3

import logging
from argparse import ArgumentParser
import time
from io import StringIO
from tools.gladier_flow.growth_curve_gladier_flow import c2_flow
from pathlib import Path
from tools.hudson_solo_auxillary.hso_functions import package_hso
from tools.hudson_solo_auxillary import solo_multi_step1, solo_multi_step2, solo_multi_step3
import pandas as pd 
import pathlib
import openpyxl
import numpy as np
import os
import datetime
import random
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from tools.ai_model import ai_actions
import scipy.stats as stats

from rpl_wei import Experiment

#from rpl_wei.wei_workcell_base import WEI

EXPERIMENT_RUN_DATAFRAMES = []
CULTURE_PAYLOAD = []
MEDIA_PAYLOAD = []
HIDEX_UPLOADS = []
PLATE_BARCODES = []
CREATED_COMPLETED_FILE = False
COMPLETED_FILE_NAME = ''
EXPERIMENT_FILE_PATH = ''

COMPLETE_HUDSON_SETUP_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/complete_hudson_setup.yaml'
STREAMLINED_HUDSON_SETUP_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/streamlined_hudson_setup.yaml'
SETUP_GROWTH_MEDIA_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/setup_growth_media.yaml'

CREATE_PLATE_T0_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/create_plate_T0.yaml'
READ_PLATE_T12_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/read_plate_T12.yaml'

DISPOSE_BOX_PLATE_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/dispose_box_plate.yaml'
DISPOSE_GROWTH_MEDIA_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/dispose_growth_media.yaml'

OPEN_CLOSE_HIDEX_FILE_PATH = '/home/rpl/workspace/BIO_workcell/applications/multi_growth_app/workflows/open_close_hidex.yaml'

exp = Experiment('127.0.0.1', '8000', 'Growth_Curve')
exp.register_exp() 
exp.events.log_local_compute("package_hso")

def sample_method_implementing_ai():
    ai_actions.load_model()
    path_name = str(pathlib.Path().resolve()) + "\\multi_growth_app\\completed_runs\\" + "07-20-2023 at 13.19.40 Completed Run.xlsx"
    ai_actions.train_model(path_name)
    SAMPLE_DATA_FOR_MODEL_ORIGINAL_ANTIBIOTIC_CONCENTRATION = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
    SAMPLE_DATA_FOR_MODEL_ORIGINAL_CELL_CONCENTRATION = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
    treatment_column, treatment_concentration, cell_column, cell_concentration, predictions = ai_actions.predict_experiment(3 ,SAMPLE_DATA_FOR_MODEL_ORIGINAL_ANTIBIOTIC_CONCENTRATION, SAMPLE_DATA_FOR_MODEL_ORIGINAL_CELL_CONCENTRATION)
    ai_actions.save_model()

#Data Processing Functions
def process_experimental_results():
    global HIDEX_UPLOADS
    global PLATE_BARCODES
    global CREATED_COMPLETED_FILE
    global COMPLETED_FILE_NAME
    global CULTURE_PAYLOAD
    global MEDIA_PAYLOAD

    globus_runs_df = pd.DataFrame()
    re_indexed = False
    #Here, we are uploading all of the
    iteration = 0
    for upload_id in HIDEX_UPLOADS:
        single_reading_df = read_globus_data(title_name = upload_id)
        if re_indexed != True:
            globus_runs_df = single_reading_df.reindex(columns=single_reading_df.columns, fill_value=None)
            re_indexed = True
        else:
            globus_runs_df = pd.concat([globus_runs_df, single_reading_df], ignore_index=True)
        iteration = iteration + 1

    run_ids = globus_runs_df['Plate #'].drop_duplicates().values
    run_ids = run_ids.astype(int)

    globus_runs_df['Plate #'] = globus_runs_df['Plate #'].astype(int)
    globus_runs_df.reset_index(drop=True, inplace=True)

    barcodes = []
    
    for run_id in run_ids:
        info_index = run_ids.tolist().index(run_id)
        barcodes.append(PLATE_BARCODES[info_index])

    folder_path = str(pathlib.Path().resolve()) + "/completed_runs/"
    # folder_path = str(pathlib.Path().resolve()) + "\\multi_growth_app\\completed_runs\\"

    if CREATED_COMPLETED_FILE:
        print("Completed File Name ", COMPLETED_FILE_NAME)
        path_name = folder_path + COMPLETED_FILE_NAME
        completed_workbook = openpyxl.load_workbook(path_name)

    else:
        print("Creating Excel Object")
        current_date = datetime.date.today()
        formatted_date = current_date.strftime("%m-%d-%Y")
        formatted_time = str(time.strftime("%H.%M.%S", time.localtime()))
        file_name = formatted_date + " at " + formatted_time +  " Completed Run" + ".xlsx"
        COMPLETED_FILE_NAME = file_name
        completed_workbook = openpyxl.Workbook() 
        CREATED_COMPLETED_FILE = True
        os.makedirs(folder_path, exist_ok=True)
        path_name = folder_path + COMPLETED_FILE_NAME
        completed_workbook.save(os.path.join(path_name))
        default_sheet = completed_workbook.active
        completed_workbook.remove(default_sheet)

    for i in range(0, len(run_ids)):
        sheet_name = "Run " + str(int(i + 1)) + " -- Barcode " + barcodes[i]
        current_sheet = completed_workbook.create_sheet(sheet_name)
        # Make this so that it just has T0, T + reading hour for each one
        plate_runs = globus_runs_df[(globus_runs_df['Plate #'] == run_ids[i])].copy()
        reading_hours = plate_runs['Reading Hour'].drop_duplicates().values
        reading_hours = reading_hours.astype(int)
        reading_hours = np.sort(reading_hours)
        pd_concat_arrays = []
        last_df = pd.DataFrame()
        for i in range(0, len(reading_hours)):
            specific_run_df = plate_runs[(plate_runs['Reading Hour'] == str(reading_hours[i]))].copy()
            specific_run_df.reset_index(drop=True, inplace=True)
            specific_run_df.rename(columns={'Result': 'T' + str(reading_hours[i])+ ' Result', 'Blank Adjusted Result': 'T' + str(reading_hours[i]) + ' Blank Adjusted Result'}, inplace=True)
            if i == 0:
                # Needs more than 5 columns in the Globus Portal to read -- stock concentration information is necessary
                specific_part_1 = specific_run_df.iloc[:, :5]
                specific_part_2 = specific_run_df.iloc[:, 5:]
                last_df = specific_part_2
                pd_concat_arrays.append(specific_part_1)
            else:
                specific_run_df = specific_run_df.loc[:, ['T' + str(reading_hours[i])+ ' Result', 'T' + str(reading_hours[i]) + ' Blank Adjusted Result']]
                pd_concat_arrays.append(specific_run_df)
        pd_concat_arrays.append(last_df)
        
        runs_df = pd.concat(pd_concat_arrays, axis=1)

        for row in dataframe_to_rows(runs_df, index=False, header=True):
            current_sheet.append(row)


    save_path = folder_path + COMPLETED_FILE_NAME
    completed_workbook.save(save_path)

    CULTURE_PAYLOAD = []
    MEDIA_PAYLOAD = []
    HIDEX_UPLOADS = []
    PLATE_BARCODES = []

def read_globus_data(title_name = ''):
    chrome_driver_path = str(pathlib.Path().resolve()) + "/tools/selenium_drivers/chromedriver.exe"
    # chrome_driver_path = str(pathlib.Path().resolve()) + "\\multi_growth_app\\tools\\selenium_drivers\\chromedriver.exe"
    driver = webdriver.Chrome(executable_path=chrome_driver_path)
    driver.get("https://acdc.alcf.anl.gov/sdl-bio/?q=*")
    search_bar = driver.find_element(By.ID, "search-input")
    search_query = "\"" + title_name + "\""
    search_bar.send_keys(search_query)
    search_bar.submit()

    link_element = driver.find_element(By.LINK_TEXT, title_name)
    link_element.click()

    desired_table_element = driver.find_element(By.XPATH, "(//table[@class='table table-striped table-bordered'])[2]")
    rows = desired_table_element.find_elements(By.TAG_NAME, "tr")
    table_data = []
    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        row_data = [cell.text for cell in cells]
        table_data.append(row_data)
    
    driver.quit()

    globus_df = pd.DataFrame(table_data)
    globus_df.columns = globus_df.iloc[0]
    globus_df = globus_df.iloc[1:]
    globus_df.reset_index(drop=True, inplace=True)

    return globus_df

def delete_experiment_excel_file():
    global EXPERIMENT_FILE_PATH
    os.remove(EXPERIMENT_FILE_PATH)
    print(EXPERIMENT_FILE_PATH)

def determine_payload_from_excel():
    global EXPERIMENT_FILE_PATH

    print("Run Log Starts Now")
    # folder_path = str(pathlib.Path().resolve()) + "\\multi_growth_app\\active_runs"
    folder_path = str(pathlib.Path().resolve()) + "/active_runs"
    files = os.listdir(folder_path)
    excel_files = [file for file in files if file.endswith(".xlsx")]
    sorted_files = sorted(excel_files, key=lambda x: os.path.getmtime(os.path.join(folder_path, x)))
    path_name = os.path.join(folder_path, sorted_files[0])
    EXPERIMENT_FILE_PATH = path_name
    print(path_name)
    workbook = openpyxl.load_workbook(filename=path_name)
    worksheet = workbook['Complete_Run_Layout']
    experiment_iterations = worksheet['B1'].value
    incubation_time_hours = worksheet['B2'].value
    incubation_time_seconds = incubation_time_hours * 3600
    added_items = 0
    for i in range(2,14):
        column_letter = chr(ord('@')+i)
        run_number_cell_id = column_letter + "3"
        media_type_cell_id = column_letter + "4"
        culture_type_cell_id = column_letter + "5"
        if(worksheet[run_number_cell_id].value != None and worksheet[media_type_cell_id].value != None and worksheet[culture_type_cell_id].value != None):      
            MEDIA_PAYLOAD.append(worksheet[media_type_cell_id].value)
            CULTURE_PAYLOAD.append(worksheet[culture_type_cell_id].value) 
            added_items = added_items + 1
    if(len(MEDIA_PAYLOAD) != experiment_iterations):
        experiment_iterations = len(MEDIA_PAYLOAD)
    
    print(CULTURE_PAYLOAD)
    print(MEDIA_PAYLOAD)
    
    return experiment_iterations, incubation_time_seconds

def setup_experiment_run_dataframes_from_stocks():
    global EXPERIMENT_RUN_DATAFRAMES
    
    stock_cell_dictionary = return_stock_dictionary("stock_cell_concentration.csv")
    stock_antibiotic_dictionary = return_stock_dictionary("stock_treatment_concentration.csv")

    for i in range(0, len(MEDIA_PAYLOAD)):
        global_stock_dictionary = {}

        #Setting up the keys in the dictionary with a blank array
        for key in stock_antibiotic_dictionary:
            if key != "Well":
                global_dictionary_key = "Antibiotic " + key

                global_stock_dictionary[global_dictionary_key] = []
        for key in stock_cell_dictionary:
            if key != "Well":
                global_dictionary_key = "Cell " + key
                global_stock_dictionary[global_dictionary_key] = []

        #Filling in the array for each key
        cell_column = CULTURE_PAYLOAD[i]
        antibiotic_column = MEDIA_PAYLOAD[i]
        for i in range(0,8):
            for j in range(1,13):
                for key in stock_antibiotic_dictionary: 
                    if key != "Well":
                        key_column = stock_antibiotic_dictionary[key]
                        array_index = antibiotic_column - 1 + 12 * i
                        key_value = key_column[array_index]
                        if str(key_value).lower() == "nan":
                            key_value = "None"
                        elif key == "Concentration (M)":
                            mod_six_j = j % 6
                            if mod_six_j == 0:
                                key_value = "0"
                            else:
                                power_raise = mod_six_j + 1
                                key_value = str(float(key_value)/2**power_raise)
                        global_dictionary_key = "Antibiotic " + key
                        global_stock_dictionary[global_dictionary_key].append(key_value)
                for key in stock_cell_dictionary:
                    if key != "Well":
                        key_column = stock_cell_dictionary[key]
                        array_index = cell_column - 1 + 12 * i
                        key_value = key_column[array_index]
                        if key == "Concentration (M)":
                                key_value = str(float(key_value)/10)
                        global_dictionary_key = "Cell " + key
                        global_stock_dictionary[global_dictionary_key].append(key_value)

        #Converting the dictionary into a Dataframe
        run_info_df = pd.DataFrame.from_dict(global_stock_dictionary)
        EXPERIMENT_RUN_DATAFRAMES.append(run_info_df)

def return_stock_dictionary(file_name):
    # file_path = str(pathlib.Path().resolve()) + "\\multi_growth_app\\active_runs\\stock_plate_information\\" + file_name
    file_path = str(pathlib.Path().resolve()) + "/active_runs/stock_plate_information/" + file_name
    stock_df = pd.read_csv(file_path)

    stock_dictionary = {}
    for column in stock_df.columns:
        stock_dictionary[column] = stock_df[column].tolist()

    return stock_dictionary

#BIO Workflow Functions
def run_experiment(total_iterations, incubation_time_sec): 
    iterations = 0
    removals = 0
    incubation_start_times = [] 
    print("Total Experimental Runs ", total_iterations)
    print("Current Iteration Variable: ", iterations)
    print("Incubation Time (Seconds): ", incubation_time_sec)
    # The experiment will run until all of the plates are used (indicated by iterations < total_iterations) and there are no more plates in the incubator (indicated by len(incubation_start_times) != 0)
    while(iterations < total_iterations or len(incubation_start_times) != 0):
        #Check to see if there are any more plates to run, indicated by total_iterations
        if(iterations < total_iterations):
            #Debug Log
            print("Starting Experiment ", iterations, ": Started Loop")
            #Set up the experiment based on the number of iterations passed.
            setup(iterations)
            #Calculate the ID of the plate needed for incubation based on the number of iterations that have passed
            liconic_id = iterations + 1
            #Run the experiment from the Hudson Solo step to the incubation step at a specified Liconic ID
            print("Starting T0 Reading")
            T0_Reading(liconic_id)
            print("Finished T0 Reading")
            assign_barcode()
            #Add the time of the incubation start to the array of 96 well plates that are currently in the incubator
            incubation_start_times.append(round(time.time()))
            #Since an iteration has now passed (the plate is in the incubator), increase the index of incubation iterations
            iterations = iterations + 1
            #Debug Log
            print("Completed Iterations: ", iterations, " ... Start Times: " , incubation_start_times)
            #Based on the total number of completed incubation iterations, determine what needs to be disposed of from the experimental setup.
            if(iterations % 2 == 0):
                print("Starting Disposal")
                dispose(iterations)
                print("Ending Disposal")
        # Check to see if delta current time and the time at which the well plate currently incubating longest exceeds the incubation time.
        if(round(time.time()) - incubation_start_times[0] > incubation_time_sec):
            #Debug Log
            print("Finishing Up Experiment ", removals, ": Ending Loop")
            #Calcuate the ID of the 96 well plate needed for removal from the incubator based on the number of plates that have already been removed.
            liconic_id = removals + 1
            #Complete the experiment starting from the removal of the 96 well plate at the specified incubation ID and ending at a Hidex T12 Reading.
            print("Starting T12 Reading")
            T12_Reading(liconic_id)
            print("Ending T12 Reading")
            #Remove the incubation start time of the plate that was just read from the array to update the new longest incubating well plate
            incubation_start_times.pop(0)
            #Increase the total number of removals that have now occurred.
            removals = removals + 1

def dispose(completed_iterations):
    #Set the default disposal index of the serial dilution plate to Stack 2.
    disposal_index = "Stack2"
    #To identify the LidNest location of where certain serial dilution plates must be held, divide the completed iterations by 2
    stack_type = completed_iterations/2
    #For the first two serial dilution plates, define the disposal index as LidNest 1 for the first serial dilution plate and LidNest 2 for the second serial dilution plate
    if(stack_type <= 2):
        disposal_index = "LidNest" + str(int(stack_type))

    #For the fourth serial dilution plate, define the disposal index as LidNest 3.
    #We are not defining the disposal index as LidNest 3 for the third serial dilution plate because there will already be a growth media plate on LidNest 3. This growth media plate will be removed in the subsequent setup function
    if(stack_type == 4):
        disposal_index = "LidNest3"
    #Add the disposal index to the payload
    payload={
        'disposal_location':  disposal_index,    
        }
    #Run the despose Yaml File with the dedicated disposal location
    run_WEI(DISPOSE_BOX_PLATE_FILE_PATH, payload)
    if(stack_type % 3 == 0):
    #If the Stack Type is a multiple of 3 (6 complete iterations of the Hudson experiment have occurred), dispose of the growth media deep well plate
        run_WEI(DISPOSE_GROWTH_MEDIA_FILE_PATH, None)

def setup(iteration_number):
    #If currently on an even number of iterations, add a tip box, serial dilution plate, and 96 well plate to the experiment.
    if(iteration_number % 2 == 0):
        #Identify the Tip Box Position Index, so it can be refilled on the Hudson Client
        payload={
                'tip_box_position': "3",
            }
        #Run the Yaml file that outlines the setup procedure for the tip box, serial dilution plate, and 96 well plate.
        print("Starting Complete Hudson Setup")
        run_WEI(COMPLETE_HUDSON_SETUP_FILE_PATH, payload)
        print("Finished Complete Setup")
        #If currently on a number of iterations that is a factor of 6, add a growth media plate to the experiment as well
        if(iteration_number % 6 == 0):
            #Specify the LidNest index of the growth well media plate that will be added to the setup (This equation assumes that there are only two growth media plates on LidNest 2 and LidNest 3 respectively for a total of 12 runs)
            LidNest_index = 2 + iteration_number/6
            #Convert the index to a readable string
            plateCrane_readable_index = "LidNest" + str(int(LidNest_index))
            print("LidNest Being Used: ", plateCrane_readable_index)
            #Add the LidNest Index to the payload
            payload={
                'lidnest_index':  plateCrane_readable_index,
            }
            #Run the Yaml file that outlines the setup procedure for the growth media plate
            print("Starting Growth Media Setup")
            run_WEI(SETUP_GROWTH_MEDIA_FILE_PATH, payload)
            print("Finished Growth Media Setup!")
    #If currently on an even number of iterations, add a 96 well plate to the experiment.
    else: 
        #Run the Yaml file that outlines the setup procedure for ONLY the 96 well plate
        run_WEI(STREAMLINED_HUDSON_SETUP_FILE_PATH, None)

def T0_Reading(liconic_plate_id):
    plate_id = '' + str(int(liconic_plate_id))
    treatment_col_id = "col" + str(int(MEDIA_PAYLOAD[liconic_plate_id-1]))
    culture_col_id = int(CULTURE_PAYLOAD[liconic_plate_id-1])
    treatment_dilution = 1
    if liconic_plate_id % 2 == 0:
        treatment_dilution = 2
    else: 
        treatment_dilution = 1
    payload={
        'temp': 37.0, 
        'humidity': 95.0,
        'shaker_speed': 30,
        "stacker": 1, 
        "slot": 1,
        "treatment": treatment_col_id, # string of treatment name. Ex. "col1", "col2"
        "culture_column": culture_col_id,  # int of cell culture column. Ex. 1, 2, 3, etc.
        "culture_dil_column": liconic_plate_id % 12, # int of dilution column for 1:10 culture dilutions. Ex. 1, 2, 3, etc.
        "media_start_column": (2*liconic_plate_id-1)%12,  # int of column to draw media from (requires 2 columns, 1 means columns 1 and 2) Ex. 1, 3, 5, etc.
        "treatment_dil_half": treatment_dilution,  #  int of which plate half to use for treatment serial dilutions. Options are 1 or 2. 
        "incubation_plate_id" : plate_id,        
        }

    # from somewhere import create_hso? or directly the solo script
    hso_1, hso_1_lines, hso_1_basename = package_hso(solo_multi_step1.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp1.hso") 
    hso_2, hso_2_lines, hso_2_basename = package_hso(solo_multi_step2.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp2.hso")  
    hso_3, hso_3_lines, hso_3_basename = package_hso(solo_multi_step3.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp3.hso")  

    # update payload with solo hso details
    payload['hso_1'] = hso_1
    payload['hso_1_lines'] = hso_1_lines
    payload['hso_1_basename'] = hso_1_basename

    payload['hso_2'] = hso_2
    payload['hso_2_lines'] = hso_2_lines
    payload['hso_2_basename'] = hso_2_basename

    payload['hso_3'] = hso_3
    payload['hso_3_lines'] = hso_3_lines
    payload['hso_3_basename'] = hso_3_basename

    # #run Growth Create Plate
    hidex_upload_id=run_WEI(CREATE_PLATE_T0_FILE_PATH, payload, Hidex_Used=True, Plate_Number=liconic_plate_id, Experiment_Run_Dataframe = EXPERIMENT_RUN_DATAFRAMES[liconic_plate_id-1])
    HIDEX_UPLOADS.append(hidex_upload_id)

def T12_Reading(liconic_plate_id):
    plate_id = '' + str(int(liconic_plate_id))

    payload={
        'temp': 37.0, 
        'humidity': 95.0,
        'shaker_speed': 30,
        "stacker": 1, 
        "slot": 1,
        "treatment": "col1", # string of treatment name. Ex. "col1", "col2"
        "culture_column": 1,  # int of cell culture column. Ex. 1, 2, 3, etc.
        "culture_dil_column": 1, # int of dilution column for 1:10 culture dilutions. Ex. 1, 2, 3, etc.
        "media_start_column": 1,  # int of column to draw media from (requires 2 columns, 1 means columns 1 and 2) Ex. 1, 3, 5, etc.
        "treatment_dil_half": 1,  #  int of which plate half to use for treatment serial dilutions. Options are 1 or 2.
        "incubation_plate_id" : plate_id,
        }

    # from somewhere import create_hso? or directly the solo script
    hso_1, hso_1_lines, hso_1_basename = package_hso(solo_multi_step1.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp1.hso") 
    hso_2, hso_2_lines, hso_2_basename = package_hso(solo_multi_step2.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp2.hso")  
    hso_3, hso_3_lines, hso_3_basename = package_hso(solo_multi_step3.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp3.hso")  

    # update payload with solo hso details
    payload['hso_1'] = hso_1
    payload['hso_1_lines'] = hso_1_lines
    payload['hso_1_basename'] = hso_1_basename

    payload['hso_2'] = hso_2
    payload['hso_2_lines'] = hso_2_lines
    payload['hso_2_basename'] = hso_2_basename

    payload['hso_3'] = hso_3
    payload['hso_3_lines'] = hso_3_lines
    payload['hso_3_basename'] = hso_3_basename

    # # #run Growth Create Plate
    hidex_upload_id=run_WEI(READ_PLATE_T12_FILE_PATH, payload, Hidex_Used=True, Plate_Number=liconic_plate_id, Experiment_Run_Dataframe = EXPERIMENT_RUN_DATAFRAMES[liconic_plate_id-1])
    HIDEX_UPLOADS.append(hidex_upload_id)

def run_WEI(file_location, payload_class, Hidex_Used = False, Plate_Number = 0, Experiment_Run_Dataframe = None):
    flow_info = exp.run_job(Path(file_location).resolve(), payload=payload_class, simulate=False)

    flow_status = exp.query_job(flow_info["job_id"])
    while(flow_status["status"] != "finished" and flow_status["status"] != "failure"):
        flow_status = exp.query_job(flow_info["job_id"])
        time.sleep(3)

    run_info = flow_status["result"]
    run_info["run_dir"] = Path(run_info["run_dir"])
    print(run_info)

    if Hidex_Used:
        t0_reading = False
        if file_location == CREATE_PLATE_T0_FILE_PATH:
            t0_reading = True
        else:
            t0_reading = False
        hidex_file_path = run_info["hist"]["run Hidex"]["action_msg"]
        hidex_file_path = hidex_file_path.replace('\\', '/')
        hidex_file_path = hidex_file_path.replace("C:/", "/C/")
        flow_title = Path(hidex_file_path) #Path(run_info["hist"]["run_assay"]["step_response"])
        fname = flow_title.name
        flow_title = flow_title.parents[0]
        experiment_time = str(time.strftime("%H_%M_%S", time.localtime()))
        experiment_name = ''
        hour_read = ''
        #Can expand this to take the hour reading based on the time incubated for -- simply pass in the incubation time (or difference) as a parameter and pass it in as hour_read
        if t0_reading:
            experiment_name = "T0_Reading"
            hour_read = '0'
        else:
            experiment_name = "T12_Reading"
            hour_read = '12'
        exp_run_df_string = Experiment_Run_Dataframe.to_csv(index=False)
        c2_flow(exp_name = experiment_name, plate_n = str(int(Plate_Number)), time = experiment_time, local_path=flow_title, fname = fname, hour=hour_read, experiment_run_dataframe = exp_run_df_string, exp = exp)
        print("Finished Uplodaing to Globus")
        return experiment_name + '_' + str(int(Plate_Number)) + '_' + experiment_time

def assign_barcode():
    current_barcode = return_barcode()
    PLATE_BARCODES.append(current_barcode)

def return_barcode():
    barcode = ""
    for _ in range(13):
        digit = random.randint(0, 9)
        barcode += str(digit)
    return barcode

#Experiment Run
def main():
    iteration_runs, incubation_time = determine_payload_from_excel()
    setup_experiment_run_dataframes_from_stocks()
    run_experiment(iteration_runs, incubation_time)
    try:
        process_experimental_results()
    except Exception as e:
        print("An exception occurred: ", e)
    # delete_experiment_excel_file()

if __name__ == "__main__":
    #main()
    main()

#!/usr/bin/env python3